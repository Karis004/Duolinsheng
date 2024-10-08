import openpyxl
import datetime
import pickle
import re
import requests
import SendKey
import logging


def review(user_name, learn_words='5'):
    learn_words = int(learn_words)
    res = []
    wave = '~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~'  #32
    file_name = 'record-'+user_name+'.xlsx'
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    today = datetime.date.today()
    next_date_one = today + datetime.timedelta(days=1)
    today_str = today.strftime('%Y-%m-%d')
    next_date_one = next_date_one.strftime('%Y-%m-%d')

    diff = [1,1,2,3,3,5,5,10,10,20,20,20,40]
    row = 2
    new_word = 1
    new_word_list = []
    for entry in sheet.iter_rows(min_row=2,values_only=True):
        if entry[0] == None:
            break
        if entry[3] == None:
            if new_word > learn_words or row == 53:
                for word in new_word_list:
                    res.append(word[0]+'\n')
                    res.append(word[1]+'\n')
                if row == 53:
                    res.append("你已学完该单词本!")
                break
            res.append(wave+'\n'+entry[0]+'\n\n'+entry[1]+'\n'+wave+'\n')
            sheet.cell(row=row, column=3).value = today_str
            sheet.cell(row=row, column=4).value = next_date_one
            sheet.cell(row=row, column=5).value = '1'
            new_word += 1
            new_word_list.append([entry[0],entry[1]])
        elif entry[3] <= today_str:
            times = int(entry[4])
            if times % 3 == 0:
                chinese_sub = re.sub('[a-zA-Z]', '', entry[1])
                res.append(chinese_sub+'\n')
                res.append(entry[0]+'\n')
            else:
                res.append(entry[0]+'\n')
                res.append(entry[1]+'\n')
            days =  diff[times] if times <= 12 else 40
            next_date = today + datetime.timedelta(days=days)
            next_date = next_date.strftime('%Y-%m-%d')
            sheet.cell(row=row, column=4).value = next_date
            sheet.cell(row=row, column=5).value = str(times+1)
        row += 1

    days = sheet.cell(row=2, column=8).value
    last_date = sheet.cell(row=2, column=9).value
    yesterday = today - datetime.timedelta(days=1)
    yesterday_str = yesterday.strftime('%Y-%m-%d')
    if last_date != today_str:
        if last_date != yesterday_str:
            new_days = 1
        else:
            new_days = int(days)+1
        sheet.cell(row=2, column=8).value = str(new_days)
        sheet.cell(row=2, column=9).value = today_str
    else:
        new_days = days

    workbook.save(file_name)
    workbook.close()
    txt_name = 'today_words-'+user_name+'.txt'
    
    with open(txt_name, 'wb') as file:
        pickle.dump(res, file)
        
    return res, new_days

def get_review(user_name):
    file_name = 'record-'+user_name+'.xlsx'
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    today = datetime.date.today()
    today_str = today.strftime('%Y-%m-%d')
    if sheet.cell(row=2, column=9).value == today_str:
        res = []
        txt_name = 'today_words-'+user_name+'.txt'
        with open(txt_name, 'rb') as file:
            res = pickle.load(file)
        days = sheet.cell(row=2, column=8).value
    else:
        res = 'Learn it once!'
        days = '0'
    workbook.close()
    return res, days


def report_mistake(mistake_list, user_name):
    tomorrow = (datetime.date.today() + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
    file_name = 'record-'+user_name+'.xlsx'
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    search_string = mistake_list[0][:-1]
    count_list = 0
    count_row = 2
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == search_string:
            sheet.cell(row=count_row, column=4).value = tomorrow
            sheet.cell(row=count_row, column=5).value = int(row[4]) - 1
            sheet.cell(row=count_row, column=6).value = row[5] + 1
            count_list += 1
            if count_list == len(mistake_list):
                break
            search_string = mistake_list[count_list][:-1]
        count_row += 1
        
    workbook.save(file_name)
    workbook.close()


def send_reminder():
    logging.info('send_reminder function is running...')
    user_name_list = SendKey.user_name_list
    for user_name in user_name_list:
        file_name = 'record-'+user_name+'.xlsx'
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        today = datetime.date.today()
        today_str = today.strftime('%Y-%m-%d')
        if sheet.cell(row=2, column=9).value != today_str:
            url = 'https://sctapi.ftqq.com/'+user_name_list[user_name]+'.send?title=Duolinsheng%20is%20waiting%20you%20:)'
            requests.get(url)
        workbook.close()
        

def error_book(user_name):
    file_name = 'record-'+user_name+'.xlsx'
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    res = []
    error_num_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == None:
            break
        if row[5] > 0:
            res.append(row[0]+'\n')
            res.append(row[1]+'\n')
            error_num_list.append(row[5])
    if len(error_num_list) == 0:
        res = 'No error!'
    workbook.close()
    return res, error_num_list
    

            